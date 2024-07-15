import dash
from dash.dependencies import Output, Input
from openpyxl import load_workbook, Workbook
from dash import html, dcc
import os.path

n = 0

# n_slots = 4    
# slot = []
# empty_color = 'white'
# full_color = 'black'

# for i in range(n_slots+1):                     #create shelf colors dictionary
#     slot_name = 'Slot_'+str(i)
#     slot.append(slot_name)
#     Shelf_Colors = dict.fromkeys(slot, empty_color)

# for i in range(n_slots+1):                     #create shelf colors dictionary
#     slot_name = 'Slot_'+str(i)
#     slot.append(slot_name)
#     Shelf_Colors = dict.fromkeys(slot, empty_color)

# book_name = 'Shelf Data.xlsx'
# file_exists = os.path.exists(book_name)

# # def create_workbook(path):
# #     '''Create new empty Excel workbook in current directory if it does not exist.'''
# #     workbook = Workbook()
# #     sheet = workbook.active
# #     sheet["A1"] = "Location Number"
# #     sheet["B1"] = "ELN Number"
# #     sheet["C1"] = "Arrival Date"
# #     sheet["D1"] = "Expiration Date"
# #     workbook.save(path)
    
# if file_exists:           #check if Excel file exists in current path
#     print('')
#     print(f'The file {book_name} exists.')
# else:
#     print('')
#     print(f'The file {book_name} does not exist; creating new workbook.') 
#     # create_workbook(book_name)
    
    
# wb = load_workbook('Shelf Data.xlsx')    
# ws = wb['Shelf 1']

n = 0

Shelf_Colors = {
    'Slot_1_Color':'white',
    'Slot_2_Color':'white',
    'Slot_3_Color':'white',
    'Slot_4_Color':'white',
    'Slot_5_Color':'white',
    'Slot_6_Color':'white',
    'Slot_7_Color':'white',
    'Slot_8_Color':'white',
    'Slot_9_Color':'white',
    'Slot_10_Color':'white',
    'Slot_11_Color':'white',
    'Slot_12_Color':'white',
    'Slot_13_Color':'white',
    'Slot_14_Color':'white',
    'Slot_15_Color':'white',
    'Slot_16_Color':'white',
    'Slot_17_Color':'white',
    'Slot_18_Color':'white',
    'Slot_19_Color':'white',
    'Slot_20_Color':'white',
    'Slot_21_Color':'white',
    'Slot_22_Color':'white',
    'Slot_23_Color':'white',
    'Slot_24_Color':'white',
    'Slot_25_Color':'white',
    'Slot_26_Color':'white',
    'Slot_27_Color':'white',
    'Slot_28_Color':'white',
    'Slot_29_Color':'white',
    'Slot_30_Color':'white',
    'Slot_31_Color':'white',
    'Slot_32_Color':'white',
    'Slot_33_Color':'white',
    'Slot_34_Color':'white',
    'Slot_35_Color':'white',
    'Slot_36_Color':'white',
    'Slot_37_Color':'white',
    'Slot_38_Color':'white',
    'Slot_39_Color':'white',
    'Slot_40_Color':'white',
    'Slot_41_Color':'white',
    'Slot_42_Color':'white',
    'Slot_43_Color':'white',
    'Slot_44_Color':'white',
    'Slot_45_Color':'white'
    }
# wb = load_workbook('/Users/ortho/SD Prototpye/Senior Design Dashboard/Shelf Data.xlsx')
wb = load_workbook('/Users/camer/GitHub/SD_Infineum/Senior Design Dashboard/Shelf Data.xlsx')
ws = wb['Shelf 1']

## updates what is currently selected

for i in range(2,47):
    if ws.cell(row = i, column = 2).value is None:
        continue
    Shelf_Colors['Slot_' + str(i-1) + '_Color'] = 'black'

def register(app:dash):
    @app.callback(Output('sample_info', 'children'),
              [Input('Slot_1', 'n_clicks'),
              Input('Slot_2', 'n_clicks'),
              Input('Slot_3', 'n_clicks'),
              Input('Slot_4', 'n_clicks'),
              Input('Slot_5', 'n_clicks'),
              Input('Slot_6', 'n_clicks'),
              Input('Slot_7', 'n_clicks'),
              Input('Slot_8', 'n_clicks'),
              Input('Slot_9', 'n_clicks'),
              Input('Slot_10', 'n_clicks'),
              Input('Slot_11', 'n_clicks'),
              Input('Slot_12', 'n_clicks'),
              Input('Slot_13', 'n_clicks'),
              Input('Slot_14', 'n_clicks'),
              Input('Slot_15', 'n_clicks'),
              Input('Slot_16', 'n_clicks'),
              Input('Slot_17', 'n_clicks'),
              Input('Slot_18', 'n_clicks'),
              Input('Slot_19', 'n_clicks'),
              Input('Slot_20', 'n_clicks'),
              Input('Slot_21', 'n_clicks'),
              Input('Slot_22', 'n_clicks'),
              Input('Slot_23', 'n_clicks'),
              Input('Slot_24', 'n_clicks'),
              Input('Slot_25', 'n_clicks'),
              Input('Slot_26', 'n_clicks'),
              Input('Slot_27', 'n_clicks'),
              Input('Slot_28', 'n_clicks'),
              Input('Slot_29', 'n_clicks'),
              Input('Slot_30', 'n_clicks'),
              Input('Slot_31', 'n_clicks'),
              Input('Slot_32', 'n_clicks'),
              Input('Slot_33', 'n_clicks'),
              Input('Slot_34', 'n_clicks'),
              Input('Slot_35', 'n_clicks'),
              Input('Slot_36', 'n_clicks'),
              Input('Slot_37', 'n_clicks'),
              Input('Slot_38', 'n_clicks'),
              Input('Slot_39', 'n_clicks'),
              Input('Slot_40', 'n_clicks'),
              Input('Slot_41', 'n_clicks'),
              Input('Slot_42', 'n_clicks'),
              Input('Slot_43', 'n_clicks'),
              Input('Slot_44', 'n_clicks'),
              Input('Slot_45', 'n_clicks'),
              ])
    def Display_Shelves(
    Slot_1,Slot_2,Slot_3,Slot_4,Slot_5,Slot_6,Slot_7,Slot_8,Slot_9,Slot_10,
    Slot_11,Slot_12,Slot_13,Slot_14,Slot_15,Slot_16,Slot_17,Slot_18,Slot_19,Slot_20,
    Slot_21,Slot_22,Slot_23,Slot_24,Slot_25,Slot_26,Slot_27,Slot_28,Slot_29,Slot_30,
    Slot_31,Slot_32,Slot_33,Slot_34,Slot_35,Slot_36,Slot_37,Slot_38,Slot_39,Slot_40,
    Slot_41,Slot_42,Slot_43,Slot_44,Slot_45):
        global n
        update_sample_info = dash.callback_context
        current_sample = update_sample_info.triggered[0]['prop_id'].split('.')[0]

        if n == 0:
            ELN_number = 'No sample selected'
            Arrial_date = ''
            experation_date =''
            n = n + 1
        else:
            for i in range(2,47):
                if current_sample == 'Slot_' + str(i-1):
                    if Shelf_Colors['Slot_' + str(i-1) + '_Color'] == 'black':
                        ELN_number = str(ws.cell(row = i, column = 2).value)
                        Arrial_date = str(ws.cell(row = i, column = 3).value)[:-9]
                        experation_date = str(ws.cell(row = i, column = 4).value)[:-9]
                    else:
                        ELN_number = 'No sample found'
                        Arrial_date = ''
                        experation_date =''

        return html.Div([
                html.Table([
                    html.Tr([
                        html.Th(str(ws.cell(row = 1, column = 2).value)),
                        html.Th(str(ws.cell(row = 1, column = 3).value)),
                        html.Th(str(ws.cell(row = 1, column =4).value)),
                    ]),
                    html.Tr([
                        html.Th(ELN_number),
                        html.Th(Arrial_date),
                        html.Th(experation_date),
                    ]),
                ], style={'Boarder':'solid', 'text-align':'center', 'font-size':'20px', 'margin-left': 'auto', 'margin-right' : 'auto'})
            ], style={'vertical-align':'top'})